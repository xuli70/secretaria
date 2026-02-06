import os
import re
import uuid

from backend.config import settings

MAX_TEXT_CHARS = 50_000


def sanitize_filename(name: str) -> str:
    """Add UUID prefix and remove unsafe characters."""
    name = os.path.basename(name)
    name = re.sub(r'[^\w.\-]', '_', name)
    return f"{uuid.uuid4().hex[:12]}_{name}"


def classify_file(ext: str) -> str:
    """Return 'document' or 'image' based on extension."""
    if ext.lower() in settings.DOCUMENT_EXTENSIONS:
        return "document"
    return "image"


def validate_file(filename: str, size: int) -> str | None:
    """Return error message or None if valid."""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in settings.ALLOWED_EXTENSIONS:
        return f"Tipo de archivo no permitido: {ext}"
    if size > settings.MAX_UPLOAD_SIZE:
        return f"Archivo demasiado grande (max {settings.MAX_UPLOAD_SIZE // (1024*1024)} MB)"
    return None


def save_file(data: bytes, safe_name: str, file_type: str) -> str:
    """Save file to disk and return the full path."""
    subdir = "documentos" if file_type == "document" else "imagenes"
    directory = os.path.join(settings.DATA_DIR, subdir)
    os.makedirs(directory, exist_ok=True)
    filepath = os.path.join(directory, safe_name)
    with open(filepath, "wb") as f:
        f.write(data)
    return filepath


def extract_text(filepath: str, ext: str) -> str | None:
    """Extract text from document files. Returns None for images."""
    ext = ext.lower()

    if ext not in settings.DOCUMENT_EXTENSIONS:
        return None

    try:
        if ext == ".pdf":
            return _extract_pdf(filepath)
        elif ext == ".docx":
            return _extract_docx(filepath)
        elif ext == ".xlsx":
            return _extract_xlsx(filepath)
        elif ext == ".txt":
            return _extract_txt(filepath)
    except Exception as e:
        return f"[Error extrayendo texto: {e}]"

    return None


def _extract_pdf(filepath: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(filepath)
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return _cap("\n".join(parts))


def _extract_docx(filepath: str) -> str:
    from docx import Document
    doc = Document(filepath)
    parts = [p.text for p in doc.paragraphs if p.text.strip()]
    return _cap("\n".join(parts))


def _extract_xlsx(filepath: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"[Hoja: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    wb.close()
    return _cap("\n".join(parts))


def _extract_txt(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return _cap(f.read())


def _cap(text: str) -> str:
    text = text.strip()
    if not text:
        return ""
    if len(text) > MAX_TEXT_CHARS:
        return text[:MAX_TEXT_CHARS] + "\n...[texto truncado]"
    return text
